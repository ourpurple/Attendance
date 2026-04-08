from datetime import datetime
from io import BytesIO
from typing import Any, Iterable, Sequence
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def _autosize_columns(ws: Worksheet, sample_rows: int = 500) -> None:
    """Auto-size columns with a capped scan for performance."""
    max_lengths = {}
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        if row_count > sample_rows:
            break
        for idx, value in enumerate(row, start=1):
            text = '' if value is None else str(value)
            length = len(text)
            if length > max_lengths.get(idx, 0):
                max_lengths[idx] = min(length, 40)

    for idx, length in max_lengths.items():
        ws.column_dimensions[get_column_letter(idx)].width = max(10, length + 2)


def build_excel_stream(
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    filename: str,
) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))

    _autosize_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_filename = filename if filename.endswith('.xlsx') else f'{filename}.xlsx'
    ascii_fallback = 'export.xlsx'
    content_disposition = (
        f'attachment; filename="{ascii_fallback}"; '
        f"filename*=UTF-8''{quote(safe_filename)}"
    )

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': content_disposition},
    )


def fmt_dt(value: Any) -> str:
    if not value:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M')
    return str(value)


def fmt_date(value: Any) -> str:
    if not value:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value)
